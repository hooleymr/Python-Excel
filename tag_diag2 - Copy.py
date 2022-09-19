import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from pprint import pprint
from openpyxl.chart.series import DataPoint
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

import matplotlib.pyplot as plt
import numpy as np


aws_regional_services = [

    "apigateway",

    "athena",

    "cassandra",

    "ec2",

    "ecs",

    "eks",

    "elasticache",

    "elasticfilesystem",

    "elasticloadbalancing",

    "elasticmapreduce",

    "dynamodb",

    "kms",

    "lambda",    

    "mq",

    "redshift",

    "rds",

    "s3",

    "sns",

    "sqs"

]
aws_regional_services.sort()

tag_names = [

    "Morpheus Instance Id",

    "ApplicationName",

    "ApplicationOwner",

    "CostCenter",

    "Description",

    "Environment",

    "Service",

    "SupportContact",

    "ProductTower",

    "SecurityPosture",

    "ApplicationCode",

    "Domain",

    "PatchGroup",

    "RequestID",

    "BackupPlan",

    "Automation",

    "BuiltBy",

    "MigratedFrom",

    "aws-migration-project-id"


]
tag_names.sort()

report_headers = [

    "Compliant",

    "AWS Account",

    "ARN"    

]

redFill = PatternFill(start_color='FFFF0000',
                        end_color='FFFF0000',
                        fill_type='solid')

start_cell = 'D2'
end_col = 'V'
#TODO above VARIABLIZE

wb = load_workbook('workbook name here')

non_compliant_list = []
all_cells = []

#TODO Define workbook in call? VARIABLIZE
def read_flags(workbook = wb, worksheets = aws_regional_services, start_cell = start_cell, end_col = end_col, flag = redFill):
    #wb = load_workbook
    for ws_name in worksheets:
        ws = wb[ws_name]
        xsl_range = ws[start_cell:f'{end_col}{ws.max_row}'] #variablize to allow different ranges in funciton call

        for cell in xsl_range:
            for x in cell:
                col_header = ws[f'{x.column_letter}1'].value 
                wb_cell = {
                    'Worksheet': ws_name,
                    'Tag' : col_header,
                    'Resource Number' : x.row,
                    'Value' : x.value, 
                    'Coordinate' : (ws_name, coordinate_from_string(f'{x.column_letter}{x.row}'))
                    }
                all_cells.append(wb_cell)

                if x != col_header and x.fill == flag: # Default range doesn't include header row but custom range may, so should still check
                    non_compliant_list.append(wb_cell)
                    #Posibly change their fill to another color to make it easier to identify which cells contain no value versus incorrect value

    return non_compliant_list


#OPYXL Wants specific cell ranges for charts, hard to do with how spread out that data can be, possible within each worksheet but hard for entire work book
#matplotlib allows specific data sets to be visualized from collected objects in non_compliant_list, can filter for each worksheet. Overall less intensive imo

def make_charts():
    total_flags = len(non_compliant_list)
    total_cells = len(all_cells)
    no_value_coords = []
    null_values = 0

   #calc amount of empty tags
    for i in range(len(non_compliant_list)):
        if non_compliant_list[i]['Value'] == None:
            no_value_coords.append(non_compliant_list[i]['Coordinate'])
            null_values += 1
    #pprint(no_value_coords)

####################################################################################################
    #Non compliant tags vs Non compliant tags with no value
    labels = f'Tags With No Value : {null_values}', f'Other Errors : {total_flags - null_values}'
    plt.rcParams["figure.figsize"] = (10, 8)
    sizes = [null_values, total_flags - null_values]
    explode = (0.1, 0)
    fig1, ax1 = plt.subplots()
    ax1.set_title("Non Compliant Tags in Current Workbook")
    ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
        shadow=True, startangle=90)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    plt.savefig('Non_value_pie.jpeg')
    plt.show()

####################################################################################################
    #Total Tags vs Total Non Compliant Tags
    total_labels = f'Non Compliant Tags : {total_flags}', f'Remaining Tags: {total_cells - total_flags}'
    plt.rcParams["figure.figsize"] = (8, 8)
    sizes = [total_flags, total_cells - total_flags]
    explode = (0.1, 0)
    fig1, ax1 = plt.subplots()
    ax1.set_title("Compliant Versus Non Compliant in Current Account")
    ax1.pie(sizes, explode=explode, labels=total_labels, autopct='%1.1f%%',
        shadow=True, startangle=90)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    plt.savefig('Total_v_non.jpeg')
    plt.show()

# ####################################################################################################
#     #WHAT MAKES AN APP NON COMPLIANT?
#     #AT LEAST ONE NC TAG IN THE APP = ALL APPS
#     #Percentage of Apps that are compliant
#     for i in range(len(total_cells)):
#         for j in aws_regional_services:
#             if total_cells[i]['Workbook']
#                #FIXME INCREMENT THROUGH EACH WORK BOOK AND COUNT AMOUNT OF NC TAGS IN EACH

#     total_labels = f'Non Compliant Tags : {total_flags}', f'Remaining tags : {total_cells}'
#     sizes = [total_flags, total_cells - total_flags]
#     explode = (0.1, 0)
#     fig1, ax1 = plt.subplots()
#     ax1.set_title("Non Compliant Apps Versus Compliant Apps")
#     ax1.pie(sizes, explode=explode, labels=total_labels, autopct='%1.1f%%',
#         shadow=True, startangle=90)
#     ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
#     plt.savefig('Total_v_non.jpeg')
#     plt.show()

####################################################################################################
    #WORST OFFENDER TAG
    freq = {}
    sorted_nc_list = sorted(non_compliant_list, key=lambda x:x['Tag']) #MAYBE OBSOLETE
    #COUNT NUM OF NC IN EACH TAG
    for i in sorted_nc_list:
        if i['Tag'] in freq:
            freq[i['Tag']] += 1
        else:
            freq[i['Tag']] = 1
    #print(freq)
    #HORIZONTAL BAR CHART
    x_pos = [i for i, _ in enumerate(freq.items())]

    plt.rcParams["figure.figsize"] = (15, 8)
    plt.barh(x_pos, list(freq.values()), color='red')
    plt.ylabel("Tags")
    plt.xlabel("Number of Non Compliant Tags")
    plt.title("Tags ranked by amount of Non compliance instances")
    plt.yticks(x_pos, list(freq.items()))
    plt.savefig('Freq_NC_bar.jpeg')
    plt.show()

    
####################################################################################################
#PER TAG -> PERCENTAGE NON COMPLIANT
#HORZ BAR CHART?

def apply_to_wb():
    #SEND GRAPHS TO EXCEL DOC
    #NEW SEPERATE? SAME DOC? NEW WS IN WB?
    return


read_flags()
make_charts()

#TODO: CREATE MATPLOTLIB CHARTS AND INSERT THEM INTO EXCEL SHEET
#TODO: Decompose workbook from top down, find errors at ws or service level, each tag has fill/no fill or len = null or != null
# if 1 row, service not utilized

#Present total non compliance per account CHECK
#Present what percentage of apps are compliant TODO
#Present which app is worst offender TODO
#Present percent compliance per service with %/total CHECK
#Present which tags are worst offender CHECK 
#Per tag, percentage non compliant TODO
#Present which tag in worst app is worst offender TODO MAYBE
